#include <stdio.h> //legacy C coding habit.
#To open a python excel sheet, and navigate to sheet called TEST1.
#Navigate through the rows in which cell value contains FRESHCO SUPER MARKET.
#Read the value corresponding to it in column 6. This has spend value.
#Sum all of it and print value.

import os

x1 = os.getcwd()
print (x1)

from openpyxl import load_workbook
wb = load_workbook(filename = 'x1.xlsx')
sheet_ranges = wb['TEST1']
count = 0
sum = 0
for rowNum in range(2, sheet_ranges.max_row):
    val = sheet_ranges.cell(row=rowNum, column=4).value
    substring = "FRESHCO HYPER MARKET"
    fullstring = val
    if (substring in fullstring):
        count = count + 1
        sum = sum + sheet_ranges.cell(row=rowNum, column=6).value
        print (count, '.)  ',  val)

print ('Total sum is ', sum)    
        


